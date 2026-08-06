"""
update_commodities.py

Single script for the commodity price database: Aluminum (LME), Wheat and
Corn (CBOT), and Barley (NCDEX).

Two modes:

1) Daily update (default, used by GitHub Actions):
       python update_commodities.py
   Fetches the latest value for all four series and upserts one row per
   commodity into the DB.

2) One-time historical backfill (only needed if the .db is ever lost,
   corrupted, or rebuilt from scratch):
       python update_commodities.py --backfill path/to/Cost_Model.xlsx
   Loads historical series from the "BBG - Yearly" sheet up to the last
   date where values actually change (later rows are carry-forward
   template placeholders and are skipped). Weekend rows are also skipped
   for all series, to stay consistent with the daily job, which never
   writes them either (none of the sources trade on weekends).

Sources & why:
-------------
Aluminum spot   -> Westmetall (westmetall.com). Publishes daily LME
                    Aluminium Cash-Settlement, sourced from the LME.
                    (The LME's own site blocks automated requests with a
                    403, so it can't be used directly.)

Aluminum future -> TradingView's public forward-curve data API
                    (scanner.tradingview.com/futures/scan). Requests the
                    full LME Aluminium High Grade ("LME:AH") forward curve
                    and picks the contract expiring ~12 months ahead of
                    the spot date, using standard futures month codes
                    (F,G,H,J,K,M,N,Q,U,V,X,Z for Jan..Dec). E.g. any day in
                    August 2026 -> contract AHQ2027. Matches Bloomberg's
                    LA13 Comdty (generic 13th-month forward) used in the
                    historical backfill.

Wheat / Corn    -> Yahoo Finance, via yfinance (tickers ZW=F and ZC=F,
                    CBOT wheat/corn futures, USD cents/bushel). Matches
                    Bloomberg's W 1 Comdty / C 1 Comdty exactly on
                    cross-checked dates.

USD/BRL         -> Banco Central do Brasil, official PTAX rate
                    (olinda.bcb.gov.br/olinda/servico/PTAX). Free, official,
                    government API, no auth needed.

INR/BRL         -> Yahoo Finance, ticker INRBRL=X. BCB's PTAX API only
                    covers 10 currencies (not INR), so this is the direct
                    cross rate instead of a synthetic USD/INR + USD/BRL
                    computation.

IGP-M (Focus)   -> Banco Central do Brasil, "Boletim Focus" market
                    expectations survey (olinda.bcb.gov.br/olinda/servico/
                    Expectativas). For any given date, stores the median
                    market expectation for IGP-M of THAT date's calendar
                    year (e.g. any day in 2026 stores the current
                    consensus for full-year IGP-M 2026), using the
                    baseCalculo=0 series (rolling last-30-days survey --
                    the one normally meant by "the Focus number"). Rolls
                    over automatically to the next year's expectation on
                    Jan 1. Free, official, no auth needed.

                    This is the driver for the Beer Cost Breakdown's
                    "Labor + Others" component (45.6% weight) -- confirmed
                    by cross-checking the spreadsheet's historical inputs
                    against BCB's realized IGP-M (SGS series 189): they
                    match almost exactly (e.g. 2020: 23.14% in both).

Barley          -> TradingView's public single-symbol quote API
                    (scanner.tradingview.com/symbol?symbol=...), symbol
                    NCDEX:BARLEYJPR1! (NCDEX Barley futures, INR/quintal).
                    Matches Bloomberg's FU1 Comdty exactly on cross-checked
                    dates. (investing.com has the same data but blocks
                    automated/datacenter requests with a 403 -- confirmed
                    by direct testing -- so it can't be used from GitHub
                    Actions.)
"""

import argparse
import json
import re
import sqlite3
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta

DB_PATH = Path(__file__).parent / "commodities_prices.db"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

MONTH_CODES = {
    1: "F", 2: "G", 3: "H", 4: "J", 5: "K", 6: "M",
    7: "N", 8: "Q", 9: "U", 10: "V", 11: "X", 12: "Z",
}

# --- Excel backfill layout (sheet "BBG - Yearly") ---
SHEET_NAME = "BBG - Yearly"
DATE_COL = 4          # column D
USDBRL_COL = 5         # column E  (USDBRL Curncy)
ALU_FUTURE_COL = 9    # column I  (LA13 Comdty)
ALU_SPOT_COL = 10     # column J  (LOAHDY LME Comdty)
BARLEY_COL = 12        # column L  (FU1 Comdty, raw INR)
WHEAT_COL = 13         # column M  (W 1 Comdty)
CORN_COL = 15          # column O  (C 1 Comdty, Chicago)
INRBRL_COL = 16        # column P  (INRBRL Curncy)
FIRST_DATA_ROW = 4


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS aluminum_lme (
            date                 TEXT PRIMARY KEY,
            spot                 REAL,
            spot_source          TEXT,
            future                REAL,
            future_contract       TEXT,
            future_target_month   TEXT,
            future_source         TEXT,
            updated_at              TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS wheat_cbot (
            date       TEXT PRIMARY KEY,
            close      REAL,
            open       REAL,
            high       REAL,
            low        REAL,
            volume     REAL,
            source     TEXT,
            updated_at   TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS corn_cbot (
            date       TEXT PRIMARY KEY,
            close      REAL,
            open       REAL,
            high       REAL,
            low        REAL,
            volume     REAL,
            source     TEXT,
            updated_at   TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS barley_ncdex (
            date       TEXT PRIMARY KEY,
            close      REAL,
            currency   TEXT DEFAULT 'INR',
            source     TEXT,
            updated_at   TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS fx_usdbrl (
            date       TEXT PRIMARY KEY,
            rate       REAL,
            source     TEXT,
            updated_at   TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS fx_inrbrl (
            date       TEXT PRIMARY KEY,
            rate       REAL,
            source     TEXT,
            updated_at   TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS igpm_focus (
            date              TEXT NOT NULL,      -- date the expectation was reported
            reference_year    INTEGER NOT NULL,     -- which year's IGP-M this expectation is for
            expectation_median REAL,
            expectation_mean   REAL,
            source            TEXT,
            updated_at          TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (date, reference_year)
        );

        CREATE TABLE IF NOT EXISTS fx_focus (
            date              TEXT PRIMARY KEY,   -- date the expectation was reported
            reference_year    INTEGER,              -- which year's USD/BRL (end-of-period) this is for
            expectation_median REAL,
            expectation_mean   REAL,
            source            TEXT,
            updated_at          TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS fx_usdbrl_guidance (
            effective_year    INTEGER PRIMARY KEY,   -- year this guidance is FOR (not when it was given)
            rate              REAL,
            source            TEXT DEFAULT 'manual_input',
            updated_at          TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS beer_cost_inflation (
            date              TEXT NOT NULL,   -- date this estimate was computed
            effective_year    INTEGER NOT NULL, -- the P&L year this estimate is for
            aluminum_yoy      REAL,
            corn_yoy          REAL,
            barley_yoy        REAL,
            igpm_yoy          REAL,
            estimated_market_fx    REAL,        -- uses realized market-average FX (fully automatic)
            estimated_company_fx   REAL,        -- uses Ambev's disclosed FX guidance (NULL if not yet input)
            weights           TEXT,             -- e.g. 'Alu 36% / Corn 6.4% / Barley 12% / IGP-M 45.6%'
            updated_at          TEXT DEFAULT (datetime('now')),
            PRIMARY KEY (date, effective_year)
        );

        CREATE TABLE IF NOT EXISTS igpm_realized (
            year        INTEGER PRIMARY KEY,   -- calendar year
            value       REAL,                   -- compounded Jan-Dec IGP-M for that year
            source      TEXT DEFAULT 'bcb/sgs-189',
            updated_at    TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS cogs_inflation_reported (
            effective_year  INTEGER PRIMARY KEY,   -- the P&L year this actual result is for
            value           REAL,                   -- Ambev's actual reported COGS/hl YoY growth
            source          TEXT DEFAULT 'company_reported',
            updated_at        TEXT DEFAULT (datetime('now'))
        );
        """
    )
    conn.commit()


def is_weekday(d: date) -> bool:
    return d.weekday() < 5  # Mon-Fri


# ----------------------------------------------------------------------
# Daily fetchers
# ----------------------------------------------------------------------

def get_westmetall_spot() -> tuple[str, float]:
    """Returns (date_str YYYY-MM-DD, cash_settlement_price) for the most
    recent trading day published on westmetall.com."""
    year = datetime.now(timezone.utc).year
    url = f"https://www.westmetall.com/en/markdaten.php?action=table&field=LME_Al_cash&year={year}"
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")
    data_rows = []
    for tr in soup.find_all("tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all("td")]
        if len(cells) >= 3 and re.match(r"^\d{2}\.\s*\w+\s*\d{4}$", cells[0]):
            data_rows.append(cells)

    if not data_rows:
        raise RuntimeError("Westmetall: could not parse any data rows from the page")

    most_recent = data_rows[0]  # table lists most recent date first
    parsed_date = datetime.strptime(most_recent[0], "%d. %B %Y").date()
    cash_price = float(most_recent[1].replace(",", ""))
    return parsed_date.strftime("%Y-%m-%d"), cash_price


def get_tradingview_aluminum_future(spot_date: date) -> tuple[str, float]:
    """Returns (contract_symbol, close_price) for the LME Aluminium High
    Grade contract expiring ~12 months after spot_date."""
    target = spot_date + relativedelta(months=12)
    code = MONTH_CODES[target.month]
    symbol = f"AH{code}{target.year}"
    full_symbol = f"LME:{symbol}"

    url = "https://scanner.tradingview.com/futures/scan?label-product=futures-forward-curve"
    payload = {
        "columns": ["pricescale", "minmov", "minmove2", "fractional", "expiration", "close", "name", "currency"],
        "filter": [{"left": "close", "operation": "nempty"}, {"left": "expiration", "operation": "nempty"}],
        "ignore_unknown_fields": False,
        "sort": {"sortBy": "expiration", "sortOrder": "asc"},
        "markets": ["futures"],
        "index_filters": [{"name": "root", "values": ["LME:AH"]}],
    }
    headers = {
        "content-type": "text/plain;charset=UTF-8",
        "accept": "application/json",
        "user-agent": USER_AGENT,
        "referer": "https://www.tradingview.com/",
    }
    r = requests.post(url, data=json.dumps(payload), headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()

    for row in data.get("data", []):
        if row["s"] == full_symbol:
            close = row["d"][5]  # index of 'close' per the columns list above
            return symbol, float(close)

    raise RuntimeError(f"TradingView: contract {full_symbol} not found in forward curve response")


def get_yfinance_ohlc(ticker: str) -> tuple[str, dict]:
    """Returns (date_str, {close, open, high, low, volume}) for the most
    recent trading day of the given Yahoo Finance ticker."""
    import yfinance as yf  # imported here so the module is only required for this path

    hist = yf.Ticker(ticker).history(period="5d")
    if hist.empty:
        raise RuntimeError(f"yfinance: no data returned for {ticker}")

    last = hist.iloc[-1]
    last_date = hist.index[-1].strftime("%Y-%m-%d")
    return last_date, {
        "close": float(last["Close"]),
        "open": float(last["Open"]),
        "high": float(last["High"]),
        "low": float(last["Low"]),
        "volume": float(last["Volume"]),
    }


def get_bcb_usdbrl(max_days_back: int = 7) -> tuple[str, float]:
    """Returns (date_str, ptax_venda_rate) for the most recent business
    day's official USD/BRL PTAX rate from the Central Bank of Brazil.
    Tries today, then walks backward day by day (weekends/holidays have
    no quote) up to max_days_back."""
    d = datetime.now(timezone.utc).date()
    for _ in range(max_days_back):
        date_param = d.strftime("%m-%d-%Y")
        url = (
            "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
            f"CotacaoDolarDia(dataCotacao='{date_param}')?$format=json"
        )
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        data = r.json()
        values = data.get("value", [])
        if values:
            return d.strftime("%Y-%m-%d"), float(values[0]["cotacaoVenda"])
        d -= relativedelta(days=1)
    raise RuntimeError(f"BCB PTAX: no USD/BRL quote found in the last {max_days_back} days")


def get_yfinance_inrbrl() -> tuple[str, float]:
    """Returns (date_str, close) for the most recent trading day of the
    INR/BRL cross rate."""
    import yfinance as yf

    hist = yf.Ticker("INRBRL=X").history(period="5d")
    if hist.empty:
        raise RuntimeError("yfinance: no data returned for INRBRL=X")
    last = hist.iloc[-1]
    last_date = hist.index[-1].strftime("%Y-%m-%d")
    return last_date, float(last["Close"])


def fetch_and_store_igpm_realized(conn) -> int:
    """Fetches BCB SGS series 189 (IGP-M monthly % change) and stores the
    compounded Jan-Dec value for every calendar year that has all 12
    months reported. Safe to call repeatedly (upserts)."""
    url = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.189/dados?formato=json"
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    by_year = {}
    for row in r.json():
        day, month, year = row["data"].split("/")
        by_year.setdefault(int(year), []).append(float(row["valor"]))

    count = 0
    for year, vals in by_year.items():
        if len(vals) != 12:
            continue
        acc = 1.0
        for v in vals:
            acc *= (1 + v / 100)
        conn.execute(
            """INSERT INTO igpm_realized (year, value) VALUES (?, ?)
               ON CONFLICT(year) DO UPDATE SET value=excluded.value, updated_at=datetime('now')""",
            (year, acc - 1),
        )
        count += 1
    return count


def get_bcb_fx_focus_latest(year: int) -> tuple[str, float, float]:
    """Returns (date_str, median, mean) for the most recently reported
    Focus market expectation of end-of-period USD/BRL for the given
    reference year."""
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata/ExpectativasMercadoAnuais"
        f"?$filter=Indicador eq 'Câmbio' and DataReferencia eq '{year}' and baseCalculo eq 0"
        "&$orderby=Data desc&$top=1&$format=json"
    )
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    values = r.json().get("value", [])
    if not values:
        raise RuntimeError(f"BCB Focus: no Câmbio expectation found for reference year {year}")
    row = values[0]
    return row["Data"], float(row["Mediana"]), float(row["Media"])


def get_bcb_igpm_focus_latest(year: int) -> tuple[str, float, float]:
    """Returns (date_str, median, mean) for the most recently reported
    Focus market expectation of IGP-M for the given reference year."""
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata/ExpectativasMercadoAnuais"
        f"?$filter=Indicador eq 'IGP-M' and DataReferencia eq '{year}' and baseCalculo eq 0"
        "&$orderby=Data desc&$top=1&$format=json"
    )
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    values = r.json().get("value", [])
    if not values:
        raise RuntimeError(f"BCB Focus: no IGP-M expectation found for reference year {year}")
    row = values[0]
    return row["Data"], float(row["Mediana"]), float(row["Media"])


def get_tradingview_barley() -> tuple[str, float]:
    """Returns (date_str today UTC, close_price) for NCDEX Barley futures.
    NCDEX:BARLEYJPR1! is the front-month rolling contract; TradingView's
    single-symbol quote endpoint doesn't expose the trade date directly,
    so we timestamp it with today's UTC date (the endpoint always
    reflects the latest available session)."""
    url = "https://scanner.tradingview.com/symbol"
    params = {
        "symbol": "NCDEX:BARLEYJPR1!",
        "fields": "close,open,high,low,currency,volume",
        "no_404": "true",
    }
    headers = {
        "accept": "application/json",
        "user-agent": USER_AGENT,
        "referer": "https://www.tradingview.com/",
    }
    r = requests.get(url, params=params, headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()
    if not data or "close" not in data:
        raise RuntimeError(f"TradingView: no barley data returned ({data!r})")

    today = datetime.now(timezone.utc).date().strftime("%Y-%m-%d")
    return today, float(data["close"])


# ----------------------------------------------------------------------
# Upserts
# ----------------------------------------------------------------------

def upsert_aluminum(conn, trade_date, spot, spot_source, future, future_contract, future_source):
    target_month = None
    if future_contract or future is not None:
        d = datetime.strptime(trade_date, "%Y-%m-%d").date()
        target_month = (d + relativedelta(months=12)).strftime("%Y-%m")
    conn.execute(
        """
        INSERT INTO aluminum_lme
            (date, spot, spot_source, future, future_contract, future_target_month, future_source)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET
            spot = COALESCE(excluded.spot, aluminum_lme.spot),
            spot_source = COALESCE(excluded.spot_source, aluminum_lme.spot_source),
            future = COALESCE(excluded.future, aluminum_lme.future),
            future_contract = COALESCE(excluded.future_contract, aluminum_lme.future_contract),
            future_target_month = COALESCE(excluded.future_target_month, aluminum_lme.future_target_month),
            future_source = COALESCE(excluded.future_source, aluminum_lme.future_source)
        """,
        (trade_date, spot, spot_source, future, future_contract, target_month, future_source),
    )


def upsert_ohlc(conn, table, trade_date, ohlc, source):
    conn.execute(
        f"""
        INSERT INTO {table} (date, close, open, high, low, volume, source)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET
            close=excluded.close, open=excluded.open, high=excluded.high,
            low=excluded.low, volume=excluded.volume, source=excluded.source
        """,
        (trade_date, ohlc["close"], ohlc["open"], ohlc["high"], ohlc["low"], ohlc["volume"], source),
    )


def upsert_fx(conn, table, trade_date, rate, source):
    conn.execute(
        f"""
        INSERT INTO {table} (date, rate, source) VALUES (?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET rate=excluded.rate, source=excluded.source
        """,
        (trade_date, rate, source),
    )


def upsert_igpm(conn, trade_date, reference_year, median, mean, source):
    conn.execute(
        """
        INSERT INTO igpm_focus (date, reference_year, expectation_median, expectation_mean, source)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(date, reference_year) DO UPDATE SET
            expectation_median=excluded.expectation_median,
            expectation_mean=excluded.expectation_mean,
            source=excluded.source
        """,
        (trade_date, reference_year, median, mean, source),
    )


def upsert_fx_focus(conn, trade_date, reference_year, median, mean, source):
    conn.execute(
        """
        INSERT INTO fx_focus (date, reference_year, expectation_median, expectation_mean, source)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET
            reference_year=excluded.reference_year,
            expectation_median=excluded.expectation_median,
            expectation_mean=excluded.expectation_mean,
            source=excluded.source
        """,
        (trade_date, reference_year, median, mean, source),
    )


# ----------------------------------------------------------------------
# Beer Cost Inflation estimate
#
# Reproduces the "Beer Cost Inflation (BRL denominated)" line from the
# Yearly Cost Model sheet:
#
#   Estimate(effective year E) =
#       36.0% * YoY(Aluminum, E)  +
#        6.4% * YoY(Corn, E)      +
#       12.0% * YoY(Barley, E)    +
#       45.6% * IGP-M expectation for E
#
#   YoY(commodity, E) = Index(E) / Index(E-1) - 1
#   Index(E) = AvgPriceUSD(hedge year = E-1) * AvgUSDBRL(hedge year = E-1)
#
# i.e. each commodity's prior-year average USD spot price, converted to
# BRL at that SAME year's own average exchange rate (this is the sheet's
# own fallback convention for when no company FX guidance is available --
# see conversation notes / README for the validation against the original
# spreadsheet, which reproduced 4.8046% / 5.7815% exactly with the
# original company-guidance FX).
#
# Weights source: "Beer Cost Breakdown", Yearly Cost Model sheet, cols X/Y.
# ----------------------------------------------------------------------

BEER_WEIGHTS = {"aluminum": 0.36, "corn": 0.064, "barley": 0.12, "igpm": 0.456}


def _avg_year(conn, table, col, year):
    return conn.execute(f"SELECT AVG({col}) FROM {table} WHERE date LIKE ?", (f"{year}-%",)).fetchone()[0]


def _avg_barley_usd_year(conn, year):
    rows = conn.execute(
        """
        SELECT b.close, f1.rate, f2.rate FROM barley_ncdex b
        JOIN fx_usdbrl f1 ON f1.date = b.date
        JOIN fx_inrbrl f2 ON f2.date = b.date
        WHERE b.date LIKE ?
        """,
        (f"{year}-%",),
    ).fetchall()
    if not rows:
        return None
    vals = [inr / (usdbrl / inrbrl) for inr, usdbrl, inrbrl in rows]
    return sum(vals) / len(vals)


def _get_igpm_expectation(conn, year):
    row = conn.execute(
        "SELECT expectation_median FROM igpm_focus WHERE reference_year=? ORDER BY date DESC LIMIT 1",
        (year,),
    ).fetchone()
    return row[0] / 100 if row else None


def compute_beer_cost_inflation(conn, effective_year: int, fx_override: dict | None = None) -> dict:
    """Computes the weighted Beer Cost Inflation estimate for the given
    effective (P&L) year.

    fx_override: optional {year: rate} dict (e.g. Ambev's disclosed FX
    guidance for effective years E and E-1). When a year isn't in the
    dict, falls back to that year's realized market-average USD/BRL --
    this is the sheet's own fallback convention for years without
    guidance (see module notes). Pass fx_override=None to use realized
    market-average FX throughout."""
    hy, hyp = effective_year - 1, effective_year - 2
    fx_override = fx_override or {}

    def fx_for(effective_yr, hedge_yr):
        return fx_override.get(effective_yr, _avg_year(conn, "fx_usdbrl", "rate", hedge_yr))

    fx_now = fx_for(effective_year, hy)
    fx_prev = fx_for(hy, hyp)  # "effective year" for the comparison point is hy (=E-1)

    def yoy(table, col, avg_fn=_avg_year):
        now = avg_fn(conn, table, col, hy) if avg_fn is _avg_year else avg_fn(conn, hy)
        prev = avg_fn(conn, table, col, hyp) if avg_fn is _avg_year else avg_fn(conn, hyp)
        if None in (now, prev, fx_now, fx_prev):
            raise RuntimeError(f"Missing input data for {table} in {hy} or {hyp}")
        return (now * fx_now) / (prev * fx_prev) - 1

    aluminum_yoy = yoy("aluminum_lme", "spot")
    corn_yoy = yoy("corn_cbot", "close")
    barley_yoy = yoy(None, None, avg_fn=_avg_barley_usd_year)
    igpm_yoy = _get_igpm_expectation(conn, effective_year)
    if igpm_yoy is None:
        raise RuntimeError(f"No IGP-M Focus expectation found for reference year {effective_year}")

    estimated = (
        BEER_WEIGHTS["aluminum"] * aluminum_yoy
        + BEER_WEIGHTS["corn"] * corn_yoy
        + BEER_WEIGHTS["barley"] * barley_yoy
        + BEER_WEIGHTS["igpm"] * igpm_yoy
    )
    return {
        "aluminum_yoy": aluminum_yoy,
        "corn_yoy": corn_yoy,
        "barley_yoy": barley_yoy,
        "igpm_yoy": igpm_yoy,
        "estimated": estimated,
    }


def get_fx_guidance(conn) -> dict:
    """Returns {effective_year: rate} for all manually-input Ambev FX
    guidance currently in the database."""
    rows = conn.execute("SELECT effective_year, rate FROM fx_usdbrl_guidance").fetchall()
    return {year: rate for year, rate in rows}


def upsert_beer_cost_inflation(conn, trade_date, effective_year, market_result: dict, company_result: dict | None):
    conn.execute(
        """
        INSERT INTO beer_cost_inflation
            (date, effective_year, aluminum_yoy, corn_yoy, barley_yoy, igpm_yoy,
             estimated_market_fx, estimated_company_fx, weights)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(date, effective_year) DO UPDATE SET
            aluminum_yoy=excluded.aluminum_yoy, corn_yoy=excluded.corn_yoy,
            barley_yoy=excluded.barley_yoy, igpm_yoy=excluded.igpm_yoy,
            estimated_market_fx=excluded.estimated_market_fx,
            estimated_company_fx=excluded.estimated_company_fx,
            weights=excluded.weights
        """,
        (
            trade_date, effective_year,
            market_result["aluminum_yoy"], market_result["corn_yoy"],
            market_result["barley_yoy"], market_result["igpm_yoy"],
            market_result["estimated"], company_result["estimated"] if company_result else None,
            "Alu 36% / Corn 6.4% / Barley 12% / IGP-M 45.6%",
        ),
    )


def upsert_barley(conn, trade_date, close, source):
    conn.execute(
        """
        INSERT INTO barley_ncdex (date, close, currency, source)
        VALUES (?, ?, 'INR', ?)
        ON CONFLICT(date) DO UPDATE SET
            close=excluded.close, currency=excluded.currency, source=excluded.source
        """,
        (trade_date, close, source),
    )


# ----------------------------------------------------------------------
# Daily update
# ----------------------------------------------------------------------

def run_daily_update() -> int:
    conn = sqlite3.connect(DB_PATH)
    create_schema(conn)
    had_error = False

    # --- Aluminum ---
    spot_date, spot_price, future_contract, future_price = None, None, None, None
    try:
        spot_date, spot_price = get_westmetall_spot()
        print(f"[Aluminum] Spot OK: {spot_date} -> {spot_price} (westmetall)")
    except Exception as e:
        had_error = True
        print(f"[Aluminum] Spot FAILED: {e}", file=sys.stderr)

    reference_date = (
        datetime.strptime(spot_date, "%Y-%m-%d").date() if spot_date else date.today()
    )
    try:
        future_contract, future_price = get_tradingview_aluminum_future(reference_date)
        print(f"[Aluminum] Future OK: {future_contract} -> {future_price} (tradingview)")
    except Exception as e:
        had_error = True
        print(f"[Aluminum] Future FAILED: {e}", file=sys.stderr)

    if spot_date or future_contract:
        row_date = spot_date or reference_date.strftime("%Y-%m-%d")
        upsert_aluminum(
            conn, row_date,
            spot_price, "westmetall" if spot_price is not None else None,
            future_price, future_contract, "tradingview" if future_price is not None else None,
        )
        conn.commit()
        print(f"[Aluminum] Saved row for {row_date}")

    # --- Wheat ---
    try:
        d, ohlc = get_yfinance_ohlc("ZW=F")
        upsert_ohlc(conn, "wheat_cbot", d, ohlc, "yahoo/ZW=F")
        conn.commit()
        print(f"[Wheat] OK: {d} -> {ohlc['close']} (yahoo/ZW=F)")
    except Exception as e:
        had_error = True
        print(f"[Wheat] FAILED: {e}", file=sys.stderr)

    # --- Corn ---
    try:
        d, ohlc = get_yfinance_ohlc("ZC=F")
        upsert_ohlc(conn, "corn_cbot", d, ohlc, "yahoo/ZC=F")
        conn.commit()
        print(f"[Corn] OK: {d} -> {ohlc['close']} (yahoo/ZC=F)")
    except Exception as e:
        had_error = True
        print(f"[Corn] FAILED: {e}", file=sys.stderr)

    # --- Barley ---
    try:
        d, close = get_tradingview_barley()
        upsert_barley(conn, d, close, "tradingview/NCDEX:BARLEYJPR1!")
        conn.commit()
        print(f"[Barley] OK: {d} -> {close} (tradingview/NCDEX:BARLEYJPR1!)")
    except Exception as e:
        had_error = True
        print(f"[Barley] FAILED: {e}", file=sys.stderr)

    # --- USD/BRL ---
    try:
        d, rate = get_bcb_usdbrl()
        upsert_fx(conn, "fx_usdbrl", d, rate, "bcb/ptax")
        conn.commit()
        print(f"[USDBRL] OK: {d} -> {rate} (bcb/ptax)")
    except Exception as e:
        had_error = True
        print(f"[USDBRL] FAILED: {e}", file=sys.stderr)

    # --- INR/BRL ---
    try:
        d, rate = get_yfinance_inrbrl()
        upsert_fx(conn, "fx_inrbrl", d, rate, "yahoo/INRBRL=X")
        conn.commit()
        print(f"[INRBRL] OK: {d} -> {rate} (yahoo/INRBRL=X)")
    except Exception as e:
        had_error = True
        print(f"[INRBRL] FAILED: {e}", file=sys.stderr)

    # --- IGP-M (Focus) ---
    for target_year in (datetime.now(timezone.utc).year, datetime.now(timezone.utc).year + 1):
        try:
            d, median, mean = get_bcb_igpm_focus_latest(target_year)
            upsert_igpm(conn, d, target_year, median, mean, "bcb/focus")
            conn.commit()
            print(f"[IGP-M Focus {target_year}] OK: {d} (ref. {target_year}) -> median={median} mean={mean} (bcb/focus)")
        except Exception as e:
            had_error = True
            print(f"[IGP-M Focus {target_year}] FAILED: {e}", file=sys.stderr)

    # --- Câmbio (Focus) ---
    for target_year in (datetime.now(timezone.utc).year, datetime.now(timezone.utc).year + 1):
        try:
            d, median, mean = get_bcb_fx_focus_latest(target_year)
            upsert_fx_focus(conn, d, target_year, median, mean, "bcb/focus")
            conn.commit()
            print(f"[Câmbio Focus {target_year}] OK: {d} (ref. {target_year}) -> median={median} mean={mean} (bcb/focus)")
        except Exception as e:
            had_error = True
            print(f"[Câmbio Focus {target_year}] FAILED: {e}", file=sys.stderr)

    # --- IGP-M realized (annual, refreshed so the just-completed year updates) ---
    try:
        n = fetch_and_store_igpm_realized(conn)
        conn.commit()
        print(f"[IGP-M realized] OK: {n} complete calendar years refreshed")
    except Exception as e:
        had_error = True
        print(f"[IGP-M realized] FAILED: {e}", file=sys.stderr)

    # --- Beer Cost Inflation estimate (current + next effective year) ---
    guidance = get_fx_guidance(conn)
    for eff_year in (datetime.now(timezone.utc).year, datetime.now(timezone.utc).year + 1):
        try:
            market_result = compute_beer_cost_inflation(conn, eff_year)
            company_result = None
            if eff_year in guidance and (eff_year - 1) in guidance:
                company_result = compute_beer_cost_inflation(conn, eff_year, fx_override=guidance)
            upsert_beer_cost_inflation(
                conn, datetime.now(timezone.utc).strftime("%Y-%m-%d"), eff_year, market_result, company_result
            )
            conn.commit()
            company_str = f"{company_result['estimated']:.4%}" if company_result else "n/a (no guidance)"
            print(f"[Beer Cost Inflation {eff_year}] OK: market_fx={market_result['estimated']:.4%} "
                  f"company_fx={company_str}")
        except Exception as e:
            had_error = True
            print(f"[Beer Cost Inflation {eff_year}] FAILED: {e}", file=sys.stderr)

    conn.close()
    return 1 if had_error else 0


# ----------------------------------------------------------------------
# One-time historical backfill
# ----------------------------------------------------------------------

def find_last_real_data_row(ws, first_row: int, cols: list[int]) -> int:
    """Find the last row where EVERY given column still shows real,
    changing data -- i.e. the most conservative cutoff across all series.
    Different series can freeze on different days (e.g. one market closes
    for the day before another), so we take the minimum (earliest) cutoff
    among all columns to avoid including a partially-stale row where some
    series already have fresh data but others are still a carry-forward
    duplicate of the previous day."""
    last_row = ws.max_row
    for r in range(last_row, first_row, -1):
        d = ws.cell(row=r, column=DATE_COL).value
        if d is not None:
            last_row = r
            break

    cutoffs = []
    for c in cols:
        prev = None
        col_cutoff = last_row
        for r in range(last_row, first_row, -1):
            v = ws.cell(row=r, column=c).value
            if prev is not None and v != prev:
                col_cutoff = r + 1
                break
            prev = v
        cutoffs.append(col_cutoff)
    return min(cutoffs)


def backfill_igpm(conn, start_year: int, end_date: date, extra_years: int = 1) -> int:
    """Fetches the FULL daily history of Focus IGP-M expectations for each
    reference year -- not just the reports published during that year
    itself, but every forward-looking projection made in prior years too
    (Focus tracks a given year several years ahead). This lets the
    dashboard reconstruct, retrospectively, what the projection for a
    year looked like at any earlier point in time -- e.g. "what did the
    market expect IGP-M 2020 to be back in mid-2019". The 'current
    beer_cost_inflation estimate' logic elsewhere only ever reads the
    LATEST row per reference_year, so this broader history is additive
    and doesn't change any existing calculation.

    extra_years: also fetches reference years beyond end_date's own year
    (default 1), matching the daily job's "current + next year" coverage
    -- otherwise the projected year (e.g. 2027, when end_date is in 2026)
    would have zero rows until the daily job runs."""
    url = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata/ExpectativasMercadoAnuais"
    total = 0
    for year in range(start_year, end_date.year + 1 + extra_years):
        full_url = (
            f"{url}?$filter=Indicador eq 'IGP-M' and DataReferencia eq '{year}' and baseCalculo eq 0"
            "&$orderby=Data asc&$top=10000&$format=json"
        )
        r = requests.get(full_url, timeout=60)
        r.raise_for_status()
        rows = r.json().get("value", [])

        for row in rows:
            d = datetime.strptime(row["Data"], "%Y-%m-%d").date()
            if d > end_date:
                continue
            if not is_weekday(d):
                continue
            date_str = d.strftime("%Y-%m-%d")
            conn.execute(
                """INSERT INTO igpm_focus
                       (date, reference_year, expectation_median, expectation_mean, source)
                   VALUES (?, ?, ?, ?, 'bcb/focus')
                   ON CONFLICT(date, reference_year) DO UPDATE SET
                       expectation_median=excluded.expectation_median,
                       expectation_mean=excluded.expectation_mean,
                       source=excluded.source""",
                (date_str, year, float(row["Mediana"]), float(row["Media"])),
            )
            total += 1
    return total


def backfill_fx_focus(conn, start_year: int, end_date: date) -> int:
    """Fetches the full daily history of Focus end-of-period USD/BRL
    expectations, year by year, keeping for each date only the
    expectation whose reference year matches that date's own calendar
    year."""
    url = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata/ExpectativasMercadoAnuais"
    total = 0
    for year in range(start_year, end_date.year + 1):
        full_url = (
            f"{url}?$filter=Indicador eq 'Câmbio' and DataReferencia eq '{year}' and baseCalculo eq 0"
            "&$orderby=Data asc&$top=10000&$format=json"
        )
        r = requests.get(full_url, timeout=60)
        r.raise_for_status()
        rows = r.json().get("value", [])

        for row in rows:
            d = datetime.strptime(row["Data"], "%Y-%m-%d").date()
            if d.year != year:
                continue
            if d > end_date:
                continue
            if not is_weekday(d):
                continue
            date_str = d.strftime("%Y-%m-%d")
            conn.execute(
                """INSERT INTO fx_focus
                       (date, reference_year, expectation_median, expectation_mean, source)
                   VALUES (?, ?, ?, ?, 'bcb/focus')
                   ON CONFLICT(date) DO UPDATE SET
                       reference_year=excluded.reference_year,
                       expectation_median=excluded.expectation_median,
                       expectation_mean=excluded.expectation_mean,
                       source=excluded.source""",
                (date_str, year, float(row["Mediana"]), float(row["Media"])),
            )
            total += 1
    return total


def backfill_cogs_reported(conn, xlsx_path: str) -> int:
    """Loads Ambev's actual reported COGS/hl YoY growth (row 44 of the
    'Yearly Cost Model' sheet -- the 'Reported' line in the backtest
    section) for every effective year available. One-time historical
    load: this isn't a live-updating series, it's a fixed snapshot from
    the original model file."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Yearly Cost Model"]
    count = 0
    for col_idx in range(2, 30):
        col = get_column_letter(col_idx)
        year = ws[f"{col}14"].value
        reported = ws[f"{col}44"].value
        if year is None or not isinstance(reported, (int, float)):
            continue
        conn.execute(
            """INSERT INTO cogs_inflation_reported (effective_year, value) VALUES (?, ?)
               ON CONFLICT(effective_year) DO UPDATE SET value=excluded.value, updated_at=datetime('now')""",
            (int(year), float(reported)),
        )
        count += 1
    return count


def run_backfill(xlsx_path: str, end_date: date | None = None) -> int:
    import openpyxl  # only needed for this mode

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    cols = [ALU_FUTURE_COL, ALU_SPOT_COL, BARLEY_COL, WHEAT_COL, CORN_COL, USDBRL_COL, INRBRL_COL]
    last_real_row = find_last_real_data_row(ws, FIRST_DATA_ROW, cols)
    cutoff_msg = ""
    if end_date is not None:
        # Walk back further if a hard end_date cap is given.
        while (ws.cell(row=last_real_row, column=DATE_COL).value.date() > end_date
               and last_real_row > FIRST_DATA_ROW):
            last_real_row -= 1
        cutoff_msg = f" (capped at requested end_date={end_date})"
    print(f"Loading rows {FIRST_DATA_ROW}..{last_real_row}{cutoff_msg} "
          f"({ws.cell(row=FIRST_DATA_ROW, column=DATE_COL).value.date()} "
          f"to {ws.cell(row=last_real_row, column=DATE_COL).value.date()})")

    conn = sqlite3.connect(DB_PATH)
    create_schema(conn)

    counts = {"aluminum": 0, "wheat": 0, "corn": 0, "barley": 0, "usdbrl": 0, "inrbrl": 0}
    skipped_weekends = 0

    for r in range(FIRST_DATA_ROW, last_real_row + 1):
        d = ws.cell(row=r, column=DATE_COL).value
        if d is None:
            continue
        d = d.date() if hasattr(d, "date") else d
        if not is_weekday(d):
            skipped_weekends += 1
            continue
        date_str = d.strftime("%Y-%m-%d")

        alu_future = ws.cell(row=r, column=ALU_FUTURE_COL).value
        alu_spot = ws.cell(row=r, column=ALU_SPOT_COL).value
        if isinstance(alu_future, (int, float)) and isinstance(alu_spot, (int, float)):
            target_month = (d + relativedelta(months=12)).strftime("%Y-%m")
            conn.execute(
                """INSERT INTO aluminum_lme
                    (date, spot, spot_source, future, future_contract, future_target_month, future_source)
                   VALUES (?, ?, 'bloomberg_backfill', ?, NULL, ?, 'bloomberg_backfill')
                   ON CONFLICT(date) DO UPDATE SET
                       spot=excluded.spot, spot_source=excluded.spot_source,
                       future=excluded.future, future_target_month=excluded.future_target_month,
                       future_source=excluded.future_source""",
                (date_str, float(alu_spot), float(alu_future), target_month),
            )
            counts["aluminum"] += 1

        wheat = ws.cell(row=r, column=WHEAT_COL).value
        if isinstance(wheat, (int, float)):
            conn.execute(
                """INSERT INTO wheat_cbot (date, close, source) VALUES (?, ?, 'bloomberg_backfill')
                   ON CONFLICT(date) DO UPDATE SET close=excluded.close, source=excluded.source""",
                (date_str, float(wheat)),
            )
            counts["wheat"] += 1

        corn = ws.cell(row=r, column=CORN_COL).value
        if isinstance(corn, (int, float)):
            conn.execute(
                """INSERT INTO corn_cbot (date, close, source) VALUES (?, ?, 'bloomberg_backfill')
                   ON CONFLICT(date) DO UPDATE SET close=excluded.close, source=excluded.source""",
                (date_str, float(corn)),
            )
            counts["corn"] += 1

        barley = ws.cell(row=r, column=BARLEY_COL).value
        if isinstance(barley, (int, float)):
            conn.execute(
                """INSERT INTO barley_ncdex (date, close, currency, source)
                   VALUES (?, ?, 'INR', 'bloomberg_backfill')
                   ON CONFLICT(date) DO UPDATE SET close=excluded.close, source=excluded.source""",
                (date_str, float(barley)),
            )
            counts["barley"] += 1

        usdbrl = ws.cell(row=r, column=USDBRL_COL).value
        if isinstance(usdbrl, (int, float)):
            conn.execute(
                """INSERT INTO fx_usdbrl (date, rate, source) VALUES (?, ?, 'bloomberg_backfill')
                   ON CONFLICT(date) DO UPDATE SET rate=excluded.rate, source=excluded.source""",
                (date_str, float(usdbrl)),
            )
            counts["usdbrl"] += 1

        inrbrl = ws.cell(row=r, column=INRBRL_COL).value
        if isinstance(inrbrl, (int, float)):
            conn.execute(
                """INSERT INTO fx_inrbrl (date, rate, source) VALUES (?, ?, 'bloomberg_backfill')
                   ON CONFLICT(date) DO UPDATE SET rate=excluded.rate, source=excluded.source""",
                (date_str, float(inrbrl)),
            )
            counts["inrbrl"] += 1

    conn.commit()
    print(f"Backfilled (skipped {skipped_weekends} weekend rows): {counts}")

    first_year = ws.cell(row=FIRST_DATA_ROW, column=DATE_COL).value.date().year
    igpm_end = end_date or ws.cell(row=last_real_row, column=DATE_COL).value.date()
    igpm_count = backfill_igpm(conn, first_year, igpm_end)
    conn.commit()
    print(f"Backfilled igpm_focus: {igpm_count} rows ({first_year}..{igpm_end.year}, capped at {igpm_end})")

    igpm_realized_count = fetch_and_store_igpm_realized(conn)
    conn.commit()
    print(f"Backfilled igpm_realized: {igpm_realized_count} complete calendar years")

    fx_focus_count = backfill_fx_focus(conn, first_year, igpm_end)
    conn.commit()
    print(f"Backfilled fx_focus: {fx_focus_count} rows ({first_year}..{igpm_end.year}, capped at {igpm_end})")

    reported_count = backfill_cogs_reported(conn, xlsx_path)
    conn.commit()
    print(f"Backfilled cogs_inflation_reported: {reported_count} years")

    conn.close()
    return 0


# ----------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--backfill", metavar="XLSX_PATH",
        help="One-time: load historical data from the given Excel file instead of running the daily update.",
    )
    parser.add_argument(
        "--end-date", metavar="YYYY-MM-DD", default=None,
        help="Optional hard cap on the last date to backfill (only used with --backfill).",
    )
    parser.add_argument(
        "--set-fx-guidance", nargs=2, metavar=("EFFECTIVE_YEAR", "RATE"),
        help="Manually record Ambev's disclosed USD/BRL guidance for a given effective (P&L) year, "
             "e.g. --set-fx-guidance 2027 5.35",
    )
    parser.add_argument(
        "--set-cogs-reported", nargs=2, metavar=("EFFECTIVE_YEAR", "VALUE_PCT"),
        help="Manually record Ambev's actually-reported COGS/hl YoY growth for a given effective (P&L) "
             "year (as a percent, e.g. 6.11 for 6.11%%), e.g. --set-cogs-reported 2025 6.11",
    )
    args = parser.parse_args()

    if args.set_fx_guidance:
        year, rate = int(args.set_fx_guidance[0]), float(args.set_fx_guidance[1])
        conn = sqlite3.connect(DB_PATH)
        create_schema(conn)
        conn.execute(
            """INSERT INTO fx_usdbrl_guidance (effective_year, rate) VALUES (?, ?)
               ON CONFLICT(effective_year) DO UPDATE SET rate=excluded.rate, updated_at=datetime('now')""",
            (year, rate),
        )
        conn.commit()
        conn.close()
        print(f"Saved FX guidance: effective year {year} -> {rate}")
        return 0

    if args.set_cogs_reported:
        year, value_pct = int(args.set_cogs_reported[0]), float(args.set_cogs_reported[1])
        conn = sqlite3.connect(DB_PATH)
        create_schema(conn)
        conn.execute(
            """INSERT INTO cogs_inflation_reported (effective_year, value) VALUES (?, ?)
               ON CONFLICT(effective_year) DO UPDATE SET value=excluded.value, updated_at=datetime('now')""",
            (year, value_pct / 100),
        )
        conn.commit()
        conn.close()
        print(f"Saved reported COGS: effective year {year} -> {value_pct}%")
        return 0

    if args.backfill:
        end_date = datetime.strptime(args.end_date, "%Y-%m-%d").date() if args.end_date else None
        return run_backfill(args.backfill, end_date=end_date)
    return run_daily_update()


if __name__ == "__main__":
    sys.exit(main())
