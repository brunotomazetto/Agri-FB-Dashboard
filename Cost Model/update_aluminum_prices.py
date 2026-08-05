"""
update_aluminum_prices.py

Single script for the aluminum spot / 12-month-forward price database.

Two modes:

1) Daily update (default, used by GitHub Actions):
       python update_aluminum_prices.py
   Fetches the latest spot (Westmetall) and 12-month-forward-equivalent
   future (TradingView forward curve) and upserts one row into the DB.

2) One-time historical backfill (only needed if the .db is ever lost,
   corrupted, or rebuilt from scratch):
       python update_aluminum_prices.py --backfill path/to/Cost_Model.xlsx
   Loads the historical series from the "BBG - Yearly" sheet (column I =
   Aluminum Future / Bloomberg LA13 Comdty, column J = Aluminum spot /
   Bloomberg LOAHDY LME Comdty) up to the last date where values actually
   change (later rows in that sheet are carry-forward template placeholders
   and are skipped). Weekend rows (Sat/Sun) are also skipped, since the LME
   doesn't trade on weekends and the daily job never writes them either —
   this keeps the backfilled history and the ongoing daily data consistent.

Spot source:   Westmetall (westmetall.com) — publishes daily LME Aluminium
               Cash-Settlement prices, sourced from the LME. (The LME's own
               site blocks automated requests with a 403, so it can't be
               used directly for automation.)

Future source: TradingView's public forward-curve data API
               (scanner.tradingview.com). Requests the full LME Aluminium
               High Grade ("LME:AH") forward curve and picks out the
               contract whose expiry is ~12 months ahead of the spot date,
               using standard futures month codes:
                 F=Jan G=Feb H=Mar J=Apr K=May M=Jun
                 N=Jul Q=Aug U=Sep V=Oct X=Nov Z=Dec
               e.g. on any day in August 2026 -> contract AHQ2027.
               This matches Bloomberg's LA13 Comdty (generic 13th-month
               forward) used in the historical backfill.
"""

import argparse
import json
import re
import sqlite3
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta

DB_PATH = Path(__file__).parent / "aluminum_prices.db"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

MONTH_CODES = {
    1: "F", 2: "G", 3: "H", 4: "J", 5: "K", 6: "M",
    7: "N", 8: "Q", 9: "U", 10: "V", 11: "X", 12: "Z",
}

# --- Excel backfill layout ---
SHEET_NAME = "BBG - Yearly"
DATE_COL = 4      # column D
FUTURE_COL = 9    # column I
SPOT_COL = 10     # column J
FIRST_DATA_ROW = 4


def create_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS aluminum_prices (
            date                 TEXT PRIMARY KEY,   -- YYYY-MM-DD
            spot                 REAL,
            spot_source          TEXT,                -- 'westmetall' | 'bloomberg_backfill'
            future                REAL,
            future_contract       TEXT,                -- e.g. 'AHQ2027' (NULL for backfill rows,
                                                         -- since Bloomberg's series is a generic
                                                         -- roll, not tied to one dated contract code)
            future_target_month   TEXT,                -- YYYY-MM the future price refers to (date + 12mo)
            future_source         TEXT,                -- 'tradingview' | 'bloomberg_backfill'
            inserted_at            TEXT DEFAULT (datetime('now'))
        )
        """
    )
    conn.commit()


def upsert(conn: sqlite3.Connection, trade_date: str,
           spot, spot_source, future, future_contract, future_source) -> None:
    target_month = None
    if future_contract or future is not None:
        d = datetime.strptime(trade_date, "%Y-%m-%d").date()
        target_month = (d + relativedelta(months=12)).strftime("%Y-%m")

    conn.execute(
        """
        INSERT INTO aluminum_prices
            (date, spot, spot_source, future, future_contract, future_target_month, future_source)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET
            spot = COALESCE(excluded.spot, aluminum_prices.spot),
            spot_source = COALESCE(excluded.spot_source, aluminum_prices.spot_source),
            future = COALESCE(excluded.future, aluminum_prices.future),
            future_contract = COALESCE(excluded.future_contract, aluminum_prices.future_contract),
            future_target_month = COALESCE(excluded.future_target_month, aluminum_prices.future_target_month),
            future_source = COALESCE(excluded.future_source, aluminum_prices.future_source)
        """,
        (trade_date, spot, spot_source, future, future_contract, target_month, future_source),
    )
    conn.commit()


# ----------------------------------------------------------------------
# Daily update
# ----------------------------------------------------------------------

def get_westmetall_spot() -> tuple[str, float]:
    """Returns (date_str YYYY-MM-DD, cash_settlement_price) for the most
    recent trading day published on westmetall.com."""
    year = datetime.now(timezone.utc).year
    url = f"https://www.westmetall.com/en/markdaten.php?action=table&field=LME_Al_cash&year={year}"
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")
    data_rows = []
    for tr in soup.find_all("tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all("td")]
        if len(cells) >= 3 and re.match(r"^\d{2}\.\s*\w+\s*\d{4}$", cells[0]):
            data_rows.append(cells)

    if not data_rows:
        raise RuntimeError("Westmetall: could not parse any data rows from the page")

    most_recent = data_rows[0]  # table lists most recent date first
    parsed_date = datetime.strptime(most_recent[0], "%d. %B %Y").date()
    cash_price = float(most_recent[1].replace(",", ""))
    return parsed_date.strftime("%Y-%m-%d"), cash_price


def get_tradingview_future(spot_date: date) -> tuple[str, float]:
    """Returns (contract_symbol, close_price) for the LME Aluminium High
    Grade contract expiring ~12 months after spot_date."""
    target = spot_date + relativedelta(months=12)
    code = MONTH_CODES[target.month]
    symbol = f"AH{code}{target.year}"
    full_symbol = f"LME:{symbol}"

    url = "https://scanner.tradingview.com/futures/scan?label-product=futures-forward-curve"
    payload = {
        "columns": ["pricescale", "minmov", "minmove2", "fractional", "expiration", "close", "name", "currency"],
        "filter": [{"left": "close", "operation": "nempty"}, {"left": "expiration", "operation": "nempty"}],
        "ignore_unknown_fields": False,
        "sort": {"sortBy": "expiration", "sortOrder": "asc"},
        "markets": ["futures"],
        "index_filters": [{"name": "root", "values": ["LME:AH"]}],
    }
    headers = {
        "content-type": "text/plain;charset=UTF-8",
        "accept": "application/json",
        "user-agent": USER_AGENT,
        "referer": "https://www.tradingview.com/",
    }
    r = requests.post(url, data=json.dumps(payload), headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()

    for row in data.get("data", []):
        if row["s"] == full_symbol:
            close = row["d"][5]  # index of 'close' per the columns list above
            return symbol, float(close)

    raise RuntimeError(f"TradingView: contract {full_symbol} not found in forward curve response")


def run_daily_update() -> int:
    conn = sqlite3.connect(DB_PATH)
    create_schema(conn)

    had_error = False
    spot_date, spot_price = None, None
    future_contract, future_price = None, None

    try:
        spot_date, spot_price = get_westmetall_spot()
        print(f"Spot OK: {spot_date} -> {spot_price} (westmetall)")
    except Exception as e:
        had_error = True
        print(f"Spot FAILED: {e}", file=sys.stderr)

    reference_date = (
        datetime.strptime(spot_date, "%Y-%m-%d").date()
        if spot_date else date.today()
    )

    try:
        future_contract, future_price = get_tradingview_future(reference_date)
        print(f"Future OK: {future_contract} -> {future_price} (tradingview)")
    except Exception as e:
        had_error = True
        print(f"Future FAILED: {e}", file=sys.stderr)

    if spot_date or future_contract:
        row_date = spot_date or reference_date.strftime("%Y-%m-%d")
        upsert(
            conn, row_date,
            spot_price, "westmetall" if spot_price is not None else None,
            future_price, future_contract, "tradingview" if future_price is not None else None,
        )
        print(f"Saved row for {row_date}")
    else:
        print("Nothing to save (both sources failed).", file=sys.stderr)

    conn.close()
    return 1 if had_error else 0


# ----------------------------------------------------------------------
# One-time historical backfill
# ----------------------------------------------------------------------

def find_last_real_data_row(ws, first_row: int) -> int:
    """Find the last row where I/J values actually change (real data),
    skipping the trailing carry-forward placeholder rows."""
    last_row = ws.max_row
    for r in range(last_row, first_row, -1):
        d = ws.cell(row=r, column=DATE_COL).value
        if d is not None:
            last_row = r
            break

    prev_i, prev_j = None, None
    for r in range(last_row, first_row, -1):
        i_val = ws.cell(row=r, column=FUTURE_COL).value
        j_val = ws.cell(row=r, column=SPOT_COL).value
        if prev_i is not None and (i_val != prev_i or j_val != prev_j):
            return r + 1
        prev_i, prev_j = i_val, j_val
    return last_row


def run_backfill(xlsx_path: str) -> int:
    import openpyxl  # only needed for this mode

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    last_real_row = find_last_real_data_row(ws, FIRST_DATA_ROW)
    print(f"Loading rows {FIRST_DATA_ROW}..{last_real_row} "
          f"({ws.cell(row=FIRST_DATA_ROW, column=DATE_COL).value.date()} "
          f"to {ws.cell(row=last_real_row, column=DATE_COL).value.date()})")

    conn = sqlite3.connect(DB_PATH)
    create_schema(conn)
    cur = conn.cursor()

    rows_to_insert = []
    skipped_weekends = 0
    for r in range(FIRST_DATA_ROW, last_real_row + 1):
        d = ws.cell(row=r, column=DATE_COL).value
        if d is None:
            continue
        spot = ws.cell(row=r, column=SPOT_COL).value
        future = ws.cell(row=r, column=FUTURE_COL).value
        if not isinstance(spot, (int, float)) or not isinstance(future, (int, float)):
            continue  # skip error cells like '#N/A N/A'

        d = d.date() if hasattr(d, "date") else d
        if d.weekday() >= 5:  # Saturday=5, Sunday=6 -> LME doesn't trade weekends
            skipped_weekends += 1
            continue

        target_month = (d + relativedelta(months=12)).strftime("%Y-%m")
        rows_to_insert.append((
            d.strftime("%Y-%m-%d"),
            float(spot), "bloomberg_backfill",
            float(future), None, target_month, "bloomberg_backfill",
        ))

    cur.executemany(
        """
        INSERT INTO aluminum_prices
            (date, spot, spot_source, future, future_contract, future_target_month, future_source)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(date) DO UPDATE SET
            spot=excluded.spot, spot_source=excluded.spot_source,
            future=excluded.future, future_contract=excluded.future_contract,
            future_target_month=excluded.future_target_month, future_source=excluded.future_source
        """,
        rows_to_insert,
    )
    conn.commit()

    count = cur.execute("SELECT COUNT(*) FROM aluminum_prices").fetchone()[0]
    print(f"Inserted/updated {len(rows_to_insert)} weekday rows "
          f"(skipped {skipped_weekends} weekend rows). Total rows in DB: {count}")
    conn.close()
    return 0


# ----------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--backfill", metavar="XLSX_PATH",
        help="One-time: load historical data from the given Excel file instead of running the daily update.",
    )
    args = parser.parse_args()

    if args.backfill:
        return run_backfill(args.backfill)
    return run_daily_update()


if __name__ == "__main__":
    sys.exit(main())
